"""
Views untuk app core (authentication & dashboard).
"""
import hashlib

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from django.shortcuts import render, redirect
from django.utils import timezone
from django.views.decorators.http import require_POST, require_http_methods
from django.views.decorators.cache import never_cache

from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.db.models import Avg, Count
from .models import SurveiVMTS


from .forms import LoginForm, MathCaptcha
from .models import LoginAttempt, DeviceSession


# ============================================
# HELPERS
# ============================================

def get_client_ip(request):
    """Ambil IP address asli client."""
    xff = request.META.get("HTTP_X_FORWARDED_FOR")
    if xff:
        return xff.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR", "0.0.0.0")


def get_user_agent(request):
    """Ambil user-agent string."""
    return request.META.get("HTTP_USER_AGENT", "")[:500]


def make_device_fingerprint(ip, user_agent):
    """Generate fingerprint dari IP prefix + user-agent."""
    ip_prefix = ".".join(ip.split(".")[:2]) if "." in ip else ip[:8]
    raw = f"{ip_prefix}|{user_agent}"
    return hashlib.sha256(raw.encode()).hexdigest()[:64]


def parse_user_agent(ua):
    """Parse user-agent sederhana ke browser, OS, device type."""
    ua_lower = ua.lower()

    if "edg/" in ua_lower:
        browser = "Edge"
    elif "chrome" in ua_lower and "safari" in ua_lower:
        browser = "Chrome"
    elif "firefox" in ua_lower:
        browser = "Firefox"
    elif "safari" in ua_lower:
        browser = "Safari"
    elif "opera" in ua_lower or "opr/" in ua_lower:
        browser = "Opera"
    else:
        browser = "Lainnya"

    if "windows" in ua_lower:
        os_name = "Windows"
    elif "mac os" in ua_lower or "macintosh" in ua_lower:
        os_name = "macOS"
    elif "android" in ua_lower:
        os_name = "Android"
    elif "iphone" in ua_lower or "ipad" in ua_lower or "ios" in ua_lower:
        os_name = "iOS"
    elif "linux" in ua_lower:
        os_name = "Linux"
    else:
        os_name = "Lainnya"

    if "mobile" in ua_lower or "android" in ua_lower or "iphone" in ua_lower:
        device_type = "Mobile"
    elif "tablet" in ua_lower or "ipad" in ua_lower:
        device_type = "Tablet"
    else:
        device_type = "Desktop"

    return browser, os_name, device_type


def count_recent_failed_attempts(ip_address, minutes=15):
    """Hitung percobaan gagal dari IP dalam X menit terakhir."""
    since = timezone.now() - timezone.timedelta(minutes=minutes)
    return LoginAttempt.objects.filter(
        ip_address=ip_address,
        status__in=[
            LoginAttempt.Status.FAILED_PASSWORD,
            LoginAttempt.Status.FAILED_USERNAME,
            LoginAttempt.Status.FAILED_CAPTCHA,
        ],
        waktu__gte=since,
    ).count()


def clear_recent_failed_attempts(ip_address, minutes=15):
    """Hapus failed attempts dari IP setelah login berhasil."""
    since = timezone.now() - timezone.timedelta(minutes=minutes)
    LoginAttempt.objects.filter(
        ip_address=ip_address,
        status__in=[
            LoginAttempt.Status.FAILED_PASSWORD,
            LoginAttempt.Status.FAILED_USERNAME,
            LoginAttempt.Status.FAILED_CAPTCHA,
        ],
        waktu__gte=since,
    ).delete()


def record_attempt(username, user, status, request):
    """Simpan login attempt ke database."""
    LoginAttempt.objects.create(
        username_attempted=username[:200],
        user=user,
        status=status,
        ip_address=get_client_ip(request),
        user_agent=get_user_agent(request),
    )


def track_device_session(user, request):
    """Track device yang login. Return (session, is_new_device)."""
    ip = get_client_ip(request)
    ua = get_user_agent(request)
    fingerprint = make_device_fingerprint(ip, ua)
    browser, os_name, device_type = parse_user_agent(ua)

    session, created = DeviceSession.objects.update_or_create(
        user=user,
        device_fingerprint=fingerprint,
        defaults={
            "ip_address": ip,
            "user_agent": ua,
            "browser_name": browser,
            "os_name": os_name,
            "device_type": device_type,
            "last_ip": ip,
            "aktif": True,
        },
    )
    return session, created


def send_new_device_notification(user, device_session, request):
    """Kirim email notifikasi login dari device baru."""
    if not user.email:
        return
    if not settings.EMAIL_HOST_USER:
        return

    subject = "Login Baru Terdeteksi — SIAKRED UNISAN"
    message = f"""Halo {user.nama_lengkap},

Ada login baru ke akun Anda di SIAKRED UNISAN dari device yang belum pernah kami kenali sebelumnya.

Detail Login:
  Waktu    : {device_session.first_seen.strftime('%d %B %Y, %H:%M:%S')} WITA
  IP       : {device_session.ip_address}
  Browser  : {device_session.browser_name}
  OS       : {device_session.os_name}
  Device   : {device_session.device_type}

Jika ini BUKAN Anda, segera:
1. Ganti password Anda
2. Hubungi administrator Pustikom UNISAN

Jika ini Anda, abaikan pesan ini.

--
SIAKRED UNISAN
Sistem Informasi Arsip Akreditasi
https://akreditasi.unisan-g.id
"""
    try:
        send_mail(
            subject=subject,
            message=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[user.email],
            fail_silently=True,
        )
    except Exception:
        pass


# ============================================
# LANDING PAGE PUBLIK
# ============================================



# ============================================================
# SUB-BATCH 12A: LANDING PUBLIK HELPERS (SIMDA-based)
# ============================================================

def _get_publik_stats():
    """Query real-time stats untuk landing publik.
    
    Sources:
    - Total prodi: SIAKRED.mapping_prodi_instrumen (prodi yang dikelola SIAKRED)
    - Total fakultas: SIMDA.master.fakultas (source of truth)
    - Total instrumen: SIAKRED
    - Total dokumen publik + sesi aktif: SIAKRED
    """
    from master_akreditasi.models import Instrumen, MappingProdiInstrumen
    from dokumen.models import Dokumen
    from sesi.models import SesiAkreditasi
    from django.db import connection
    
    # Total prodi: dari mapping SIAKRED
    total_prodi = MappingProdiInstrumen.objects.filter(aktif=True).count()
    
    # Total fakultas: dari SIMDA (source of truth)
    total_fakultas = 0
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT COUNT(*) FROM master.fakultas WHERE LOWER(status) = 'aktif'"
            )
            total_fakultas = cursor.fetchone()[0]
    except Exception:
        # Fallback: hitung dari prefix kode prodi
        prodi_codes = list(
            MappingProdiInstrumen.objects.filter(aktif=True)
            .values_list('kode_prodi', flat=True)
        )
        total_fakultas = len(set(c[0].upper() for c in prodi_codes if c))
    
    # Instrumen & dokumen
    total_instrumen = Instrumen.objects.filter(aktif=True).count()
    total_dokumen_publik = 0
    try:
        total_dokumen_publik = Dokumen.objects.filter(
            status='FINAL',
            status_akses='TERBUKA',
        ).count()
    except Exception:
        pass
    
    # Sesi aktif
    total_sesi_aktif = SesiAkreditasi.objects.exclude(
        status__in=['SELESAI', 'DIBATALKAN']
    ).count()
    
    return {
        'total_prodi': total_prodi,
        'total_instrumen': total_instrumen,
        'total_fakultas': total_fakultas,
        'total_dokumen_publik': total_dokumen_publik,
        'total_sesi_aktif': total_sesi_aktif,
    }


def _get_fakultas_list():
    """Get list fakultas dari SIMDA + prodi-nya dari SIAKRED mapping.
    
    SOURCE OF TRUTH:
    - Fakultas list: SIMDA.master.fakultas
    - Prodi per fakultas: SIMDA.master.program_studi JOIN SIAKRED.mapping_prodi_instrumen
    
    Returns: list of dict {kode, nama, nama_singkat, prodi_list: [...]}
    """
    from master_akreditasi.models import MappingProdiInstrumen
    from django.db import connection
    
    # 1. Ambil semua fakultas dari SIMDA
    fakultas_dict = {}
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT kode_fakultas, nama_fakultas, nama_singkat, urutan
                FROM master.fakultas
                WHERE LOWER(status) = 'aktif'
                ORDER BY urutan NULLS LAST, kode_fakultas
            """)
            for row in cursor.fetchall():
                kode = row[0]
                fakultas_dict[kode] = {
                    'kode': kode,
                    'nama': row[1],
                    'nama_singkat': row[2] or kode,
                    'urutan': row[3] or 99,
                    'prodi_list': [],
                }
    except Exception:
        # Fallback kalau SIMDA tidak accessible
        pass
    
    # 2. Ambil prodi dari SIMDA dengan kode_fakultas-nya
    prodi_to_fakultas = {}
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT kode_prodi, kode_fakultas FROM master.program_studi"
            )
            for row in cursor.fetchall():
                prodi_to_fakultas[row[0]] = row[1]
    except Exception:
        pass
    
    # 3. Loop mapping SIAKRED, assign prodi ke fakultas yang benar (dari SIMDA)
    mapping_qs = MappingProdiInstrumen.objects.filter(
        aktif=True
    ).select_related('instrumen').order_by('kode_prodi')
    
    for m in mapping_qs:
        if not m.kode_prodi:
            continue
        
        # Cari kode_fakultas asli dari SIMDA
        kode_fak = prodi_to_fakultas.get(m.kode_prodi)
        
        # Fallback: derivation dari huruf pertama
        if not kode_fak:
            prefix = m.kode_prodi[0].upper()
            _FALLBACK_MAP = {
                'T': 'FT', 'E': 'FE', 'H': 'FH',
                'K': 'FK', 'P': 'FP', 'S': 'FS',
            }
            kode_fak = _FALLBACK_MAP.get(prefix)
        
        if not kode_fak or kode_fak not in fakultas_dict:
            continue
        
        instrumen_nama = (
            getattr(m.instrumen, 'nama_singkat', None)
            or getattr(m.instrumen, 'nama_resmi', None)
            or getattr(m.instrumen, 'nama', None)
            or str(m.instrumen)
        )
        
        fakultas_dict[kode_fak]['prodi_list'].append({
            'kode': m.kode_prodi,
            'nama': m.nama_prodi,
            'instrumen': instrumen_nama,
        })
    
    # 4. Enrich dengan theme warna dari FakultasTheme model
    from core.models import FakultasTheme
    theme_map = FakultasTheme.get_theme_map()
    
    DEFAULT_THEME = {"primary": "#2563EB", "light": "#DBEAFE", "icon": "graduation-cap"}
    for kode, fak in fakultas_dict.items():
        theme = theme_map.get(kode, DEFAULT_THEME)
        fak["warna_primary"] = theme["primary"]
        fak["warna_light"] = theme["light"]
        fak["icon_nama"] = theme["icon"]
    
    # 5. Return sorted list (by urutan SIMDA)
    result = sorted(fakultas_dict.values(), key=lambda x: x.get('urutan', 99))
    return result


def landing_page(request):
    """Landing publik SIAKRED. Hybrid: single page scroll dengan section lengkap."""
    from core.models import SiteProfile
    site = SiteProfile.get_instance()
    profile = SiteProfile.get_instance()
    stats = _get_publik_stats()
    fakultas_list = _get_fakultas_list()
    
    # Ambil preview dokumen publik terbaru (6 dokumen)
    from dokumen.models import Dokumen
    dokumen_terbaru = Dokumen.objects.filter(
        status='FINAL',
        status_akses='TERBUKA',
    ).select_related('butir_dokumen').order_by('-tanggal_dibuat')[:6]
    
    try:
        from master_akreditasi.simda import get_institusi_from_simda
        institusi_simda = get_institusi_from_simda() or {}
    except Exception:
        institusi_simda = {}
    
    # Ensure all expected keys exist (prevent template crash)
    default_keys = ['akreditasi', 'no_sk_akreditasi', 'npsn', 'nama_resmi', 'nama_singkat', 
                    'alamat', 'telepon', 'email', 'website', 'tgl_sk_akreditasi', 
                    'berlaku_sampai', 'tgl_berdiri', 'visi', 'misi', 'tujuan']
    for k in default_keys:
        institusi_simda.setdefault(k, "")
    
    context = {
        'profile': profile,
        'institusi_simda': institusi_simda,
        'stats': stats,
        'fakultas_list': fakultas_list,
        'dokumen_terbaru': dokumen_terbaru,
        'site': site,
    }
    return render(request, 'landing/landing.html', context)

def login_view(request):
    """Login dengan CAPTCHA, lockout, device tracking."""

    if request.user.is_authenticated:
        return redirect("core:dashboard")

    ip = get_client_ip(request)
    failed_count = count_recent_failed_attempts(ip, minutes=15)
    require_captcha = failed_count >= 2

    if failed_count >= 5:
        return render(
            request,
            "auth/locked.html",
            {
                "page_title": "Akun Dikunci Sementara",
                "wait_minutes": 15,
            },
            status=429,
        )

    if request.method == "POST":
        form = LoginForm(
            request.POST,
            request=request,
            require_captcha=require_captcha,
        )

        identifier = request.POST.get("identifier", "").strip()

        if form.is_valid():
            user = form.get_user()
            remember_me = form.cleaned_data.get("remember_me", False)

            if not remember_me:
                request.session.set_expiry(0)
            else:
                request.session.set_expiry(60 * 60 * 24 * 14)

            device_session, is_new_device = track_device_session(user, request)

            if is_new_device:
                send_new_device_notification(user, device_session, request)

            # Reset failed attempts saat login sukses
            clear_recent_failed_attempts(ip, minutes=15)

            login(request, user)
            record_attempt(identifier, user, LoginAttempt.Status.SUCCESS, request)

            messages.success(
                request,
                f"Selamat datang, {user.nama_lengkap}!"
            )
            return redirect("core:dashboard")
        else:
            error_codes = []
            for errs in form.errors.as_data().values():
                for e in errs:
                    error_codes.append(e.code)

            if "captcha_required" in error_codes or "captcha_invalid" in error_codes:
                status = LoginAttempt.Status.FAILED_CAPTCHA
            elif "inactive" in error_codes:
                status = LoginAttempt.Status.FAILED_INACTIVE
            else:
                status = LoginAttempt.Status.FAILED_PASSWORD

            record_attempt(identifier or "(kosong)", None, status, request)

            failed_count = count_recent_failed_attempts(ip, minutes=15)
            require_captcha = failed_count >= 2
    else:
        form = LoginForm(request=request, require_captcha=require_captcha)

    captcha_data = None
    if require_captcha:
        question, answer, token = MathCaptcha.generate()
        captcha_data = {
            "question": question,
            "token": token,
        }
        if "captcha_question" in form.fields:
            form.fields["captcha_question"].initial = question
            form.fields["captcha_token"].initial = token

    context = {
        "page_title": "Login — SIAKRED UNISAN",
        "form": form,
        "require_captcha": require_captcha,
        "captcha": captcha_data,
        "failed_count": failed_count,
    }
    return render(request, "auth/login.html", context)


# ============================================
# LOGOUT VIEW
# ============================================

@require_POST
def logout_view(request):
    """Logout & redirect ke landing."""
    if request.user.is_authenticated:
        logout(request)
        messages.info(request, "Anda telah berhasil logout.")
    return redirect("core:landing")


# ============================================
# DASHBOARD
# ============================================

@login_required(login_url="/login/")
def dashboard_view(request):
    """Dashboard internal SIAKRED (placeholder)."""
    user = request.user
    scopes = user.scopes.filter(aktif=True) if hasattr(user, "scopes") else []

    current_fingerprint = make_device_fingerprint(
        get_client_ip(request), get_user_agent(request)
    )
    last_device = (
        DeviceSession.objects.filter(user=user)
        .exclude(device_fingerprint=current_fingerprint)
        .order_by("-last_seen")
        .first()
    )

    context = {
        "page_title": "Dashboard",
        "active_menu": "dashboard",
        "user": user,
        "scopes": scopes,
        "last_device": last_device,
    }
    return render(request, "dashboard/index.html", context)


# ============================================
# NOTIFIKASI VIEWS (Step 9H.3)
# ============================================

from django.core.paginator import Paginator
from django.utils import timezone as _tz
from django.views.decorators.http import require_POST


@login_required(login_url='/login/')
def notifikasi_list(request):
    """Halaman list semua notifikasi user."""
    from core.models import Notifikasi
    
    filter_by = request.GET.get('filter', 'all')
    base_qs = Notifikasi.objects.filter(penerima=request.user)
    
    if filter_by == 'unread':
        qs = base_qs.filter(sudah_dibaca=False)
    elif filter_by == 'read':
        qs = base_qs.filter(sudah_dibaca=True)
    else:
        qs = base_qs
    
    qs = qs.select_related('dibuat_oleh', 'dokumen').order_by('-tanggal_dibuat')
    
    paginator = Paginator(qs, 20)
    page_num = request.GET.get('page', 1)
    page = paginator.get_page(page_num)
    
    context = {
        'page_title': 'Notifikasi',
        'active_menu': 'notifikasi',
        'page': page,
        'filter_by': filter_by,
        'total_count': base_qs.count(),
        'unread_count': base_qs.filter(sudah_dibaca=False).count(),
    }
    return render(request, 'core/notifikasi_list.html', context)


@login_required(login_url='/login/')
def notifikasi_read(request, notif_id):
    """Mark notifikasi sebagai dibaca + redirect ke url_action."""
    from core.models import Notifikasi
    from django.shortcuts import get_object_or_404
    
    notif = get_object_or_404(Notifikasi, pk=notif_id, penerima=request.user)
    notif.tandai_dibaca()
    
    # Redirect ke url_action kalau ada, kalau tidak ke list
    target = notif.url_action or '/notifikasi/'
    return redirect(target)


@login_required(login_url='/login/')
@require_POST
def notifikasi_mark_all_read(request):
    """Mark semua notifikasi user sebagai dibaca."""
    from core.models import Notifikasi
    
    updated = Notifikasi.objects.filter(
        penerima=request.user,
        sudah_dibaca=False,
    ).update(
        sudah_dibaca=True,
        tanggal_dibaca=_tz.now(),
    )
    
    messages.success(request, f'{updated} notifikasi ditandai dibaca.')
    return redirect('core:notifikasi_list')


# ============================================
# HELP CENTER (Step User Manual)
# ============================================

@login_required(login_url='/login/')
def help_index(request):
    """Halaman index pusat bantuan dengan kategori per role."""
    user = request.user
    
    # Detect role user untuk highlight section yang relevan
    user_roles = []
    if user.is_superuser:
        user_roles.append('admin')
    if hasattr(user, 'scopes'):
        for scope in user.scopes.filter(aktif=True):
            lvl = (scope.level or '').upper()
            if lvl in ('UNIVERSITAS', 'REKTORAT', 'SUPER', 'ADMIN'):
                user_roles.append('admin')
            elif lvl in ('BIRO', 'LP3M', 'LP2M'):
                user_roles.append('lp3m')
            elif lvl == 'FAKULTAS':
                user_roles.append('dekan')
            elif lvl == 'PRODI':
                user_roles.append('kaprodi')
    
    context = {
        'page_title': 'Pusat Bantuan',
        'active_menu': 'help',
        'user_roles': list(set(user_roles)),
    }
    return render(request, 'core/help_index.html', context)


@login_required(login_url='/login/')
def help_section(request, section):
    """Halaman detail per section: 'uploader', 'verifikator', 'admin', 'faq'."""
    valid_sections = {'uploader', 'verifikator', 'admin', 'faq'}
    if section not in valid_sections:
        return redirect('core:help_index')
    
    section_titles = {
        'uploader': 'Panduan untuk Kaprodi/Dosen',
        'verifikator': 'Panduan untuk LP3M/Dekan',
        'admin': 'Panduan untuk Admin/Rektorat',
        'faq': 'FAQ & Troubleshooting',
    }
    
    context = {
        'page_title': section_titles.get(section, 'Pusat Bantuan'),
        'active_menu': 'help',
        'section': section,
        'section_title': section_titles.get(section),
    }
    return render(request, f'core/help_{section}.html', context)


# ============================================================
# SUB-BATCH 12A: LANDING PUBLIK DETAIL PAGES (placeholder)
# ============================================================

def fakultas_list(request):
    """List semua fakultas (placeholder, implementasi di 12E)."""
    fakultas_list = _get_fakultas_list()
    context = {
        'page_title': 'Daftar Fakultas',
        'fakultas_list': fakultas_list,
    }
    return render(request, 'landing/fakultas_list.html', context)


def fakultas_detail(request, kode):
    """Detail 1 fakultas + list prodi (enriched dari SIMDA + profile)."""
    from django.http import Http404
    from django.db import connection
    from core.models import FakultasProfile
    
    fakultas_list = _get_fakultas_list()
    fakultas = next((f for f in fakultas_list if f['kode'].upper() == kode.upper()), None)
    
    if not fakultas:
        raise Http404(f'Fakultas dengan kode {kode} tidak ditemukan')
    
    # Enrich prodi_list dengan data SIMDA
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT kode_prodi, jenjang, akreditasi, email_prodi
                FROM master.program_studi
                WHERE UPPER(kode_fakultas) = UPPER(%s)
            """, [fakultas['kode']])
            simda_data = {r[0]: {'jenjang': r[1], 'akreditasi': r[2], 'email_prodi': r[3]} for r in cursor.fetchall()}
        
        for p in fakultas['prodi_list']:
            extra = simda_data.get(p['kode'], {})
            p['jenjang'] = extra.get('jenjang', '-')
            p['akreditasi'] = extra.get('akreditasi', '-')
            p['email_prodi'] = extra.get('email_prodi', '')
    except Exception:
        pass
    
    # Ambil FakultasProfile (foto dekan, visi misi override)
    fakultas_profile = FakultasProfile.objects.filter(
        kode_fakultas=fakultas['kode'], aktif=True
    ).first()
    
    context = {
        'page_title': fakultas['nama'],
        'fakultas': fakultas,
        'fakultas_profile': fakultas_profile,
    }
    return render(request, 'landing/fakultas_detail.html', context)


def prodi_detail(request, kode):
    """Detail 1 prodi (enriched dari SIMDA + theme warna + profile)."""
    from django.http import Http404
    from master_akreditasi.models import MappingProdiInstrumen
    from master_akreditasi.simda import get_institusi_from_simda, get_prodi_detail_from_simda
    from core.models import FakultasTheme, FakultasProfile, ProdiProfile
    
    kode_upper = kode.upper()
    
    try:
        mapping = MappingProdiInstrumen.objects.select_related('instrumen').get(
            kode_prodi=kode_upper,
            aktif=True,
        )
    except MappingProdiInstrumen.DoesNotExist:
        raise Http404(f'Prodi dengan kode {kode} tidak ditemukan')
    
    # Ambil detail dari SIMDA
    simda_data = get_prodi_detail_from_simda(kode_upper) or {}
    
    # Ambil theme warna fakultas
    kode_fakultas = simda_data.get('kode_fakultas', '')
    theme_obj = None
    if kode_fakultas:
        theme_obj = FakultasTheme.objects.filter(
            kode_fakultas=kode_fakultas, aktif=True
        ).first()
    
    theme_data = {
        'primary': theme_obj.warna_primary if theme_obj else '#2563EB',
        'light': theme_obj.warna_light if theme_obj else '#DBEAFE',
        'icon': theme_obj.icon_nama if theme_obj else 'graduation-cap',
    }
    
    # Ambil ProdiProfile (foto kaprodi, visi misi)
    prodi_profile = ProdiProfile.objects.filter(kode_prodi=kode_upper, aktif=True).first()
    
    # Ambil FakultasProfile (untuk nama dekan, dll)
    fakultas_profile = None
    if kode_fakultas:
        fakultas_profile = FakultasProfile.objects.filter(
            kode_fakultas=kode_fakultas, aktif=True
        ).first()
    
    # Sesi akreditasi terbaru
    from sesi.models import SesiAkreditasi
    sesi_terbaru = SesiAkreditasi.objects.filter(
        kode_prodi=kode_upper
    ).order_by('-tanggal_mulai').first()
    
    context = {
        'page_title': mapping.nama_prodi,
        'prodi': mapping,
        'simda': simda_data,
        'theme': theme_data,
        'kode_fakultas': kode_fakultas,
        'prodi_profile': prodi_profile,
        'fakultas_profile': fakultas_profile,
        'sesi_terbaru': sesi_terbaru,
    }
    return render(request, 'landing/prodi_detail.html', context)


from master_akreditasi.models import MappingProdiInstrumen

def _get_survei_context(error=None):
    """Helper untuk build context survei dengan daftar prodi."""
    from core.models import SiteProfile
    from django.db import connections
    site = SiteProfile.get_instance()
    cursor = connections['default'].cursor()
    cursor.execute("""
        SELECT m.kode_prodi, m.nama_prodi, p.jenjang
        FROM mapping_prodi_instrumen m
        LEFT JOIN master.program_studi p ON p.kode_prodi = m.kode_prodi
        WHERE m.aktif = true
        ORDER BY m.nama_prodi, p.jenjang
    """)
    rows = cursor.fetchall()
    daftar_prodi = [
        {
            'kode_prodi': row[0],
            'nama_prodi': row[1],
            'jenjang': row[2] or '',
            'label': f"{row[1]} ({row[2]})" if row[2] else row[1],
        }
        for row in rows
    ]
    return {
        'stats': SurveiVMTS.objects.aggregate(
            total=Count('id'),
            avg_v=Avg('skor_v'),
            avg_m=Avg('skor_m'),
            avg_t=Avg('skor_t'),
            avg_s=Avg('skor_s'),
            avg_total=Avg('skor_total'),
        ),
        'target_responden': site.survei_vmts_target,
        'tahun_akademik': site.survei_vmts_tahun_akademik,
        'daftar_prodi': daftar_prodi,
        'site': site,
        'error': error,
    }


def survei_vmts(request):
    from core.models import SiteProfile
    site = SiteProfile.get_instance()
    
    # Cek apakah survei sedang aktif
    if not site.survei_vmts_aktif:
        return render(request, 'survei/vmts_tutup.html', {
            'tahun_akademik': site.survei_vmts_tahun_akademik,
        })
    
    return render(request, 'survei/vmts_form.html', _get_survei_context())


def kirim_vmts(request):
    if request.method != 'POST':
        return redirect('core:survei_vmts')

    try:
        nama     = request.POST.get('nama', '').strip()
        status   = request.POST.get('status', '')
        prodi    = request.POST.get('prodi', '').strip()
        angkatan = request.POST.get('angkatan', '').strip()

        if not status or not prodi:
            ctx = _get_survei_context(error='Status dan Program Studi wajib diisi.')
            ctx['old'] = request.POST  # ← tambahkan ini
            return render(request, 'survei/vmts_form.html', ctx)

        def hitung_skor(keys):
            nilai = []
            for k in keys:
                v = request.POST.get(k)
                if v and v.isdigit():
                    nilai.append(int(v))
            return round(sum(nilai) / len(nilai), 2) if nilai else 0

        skor_v = hitung_skor(['V1', 'V2', 'V3', 'V4'])
        skor_m = hitung_skor(['M1', 'M2', 'M3', 'M4'])
        skor_t = hitung_skor(['T1', 'T2', 'T3', 'T4'])
        skor_s = hitung_skor(['S1', 'S2', 'S3', 'S4'])

        # Validasi semua pilar harus diisi
        errors = []
        if skor_v == 0: errors.append('Visi')
        if skor_m == 0: errors.append('Misi')
        if skor_t == 0: errors.append('Tujuan')
        if skor_s == 0: errors.append('Sasaran')

        if errors:
            ctx = _get_survei_context(
                error=f'Mohon lengkapi semua bagian. Belum diisi: {", ".join(errors)}.'
            )
            ctx['old'] = request.POST
            return render(request, 'survei/vmts_form.html', ctx)

        # Hitung skor total
        skor_total = round(
            (skor_v + skor_m + skor_t + skor_s) / 4, 2
        )

        media_str = ','.join(request.POST.getlist('media'))
        x_forwarded = request.META.get('HTTP_X_FORWARDED_FOR')
        ip = x_forwarded.split(',')[0] if x_forwarded else request.META.get('REMOTE_ADDR')

        SurveiVMTS.objects.create(
            nama=nama or None,
            status=status,
            prodi=prodi,
            angkatan=angkatan or None,
            skor_v=skor_v,
            skor_m=skor_m,
            skor_t=skor_t,
            skor_s=skor_s,
            skor_total=skor_total,
            media_sosialisasi=media_str or None,
            masukan=request.POST.get('masukan', '').strip() or None,
            saran=request.POST.get('saran', '').strip() or None,
            ip_address=ip,
        )

        return redirect('core:survei_vmts_sukses')

    except Exception as e:
        ctx = _get_survei_context(error=f'Terjadi kesalahan: {str(e)}')
        ctx['old'] = request.POST  # ← tambahkan ini
        return render(request, 'survei/vmts_form.html', ctx)


def survei_vmts_sukses(request):
    from core.models import SiteProfile
    site = SiteProfile.get_instance()
    return render(request, 'survei/vmts_sukses.html', {'site': site})


def dashboard_vmts(request):
    if not request.user.is_authenticated:
        return redirect('core:login')

    from core.models import SiteProfile, ProdiProfile
    site = SiteProfile.get_instance()
    prodi_filter = request.GET.get('prodi', '')

    data = SurveiVMTS.objects.all()
    if prodi_filter:
        data = data.filter(prodi=prodi_filter)

    stats = data.aggregate(
        total=Count('id'),
        avg_v=Avg('skor_v'),
        avg_m=Avg('skor_m'),
        avg_t=Avg('skor_t'),
        avg_s=Avg('skor_s'),
        avg_total=Avg('skor_total'),
    )

    per_status = data.values('status').annotate(
        jumlah=Count('id'),
        rata=Avg('skor_total')
    ).order_by('status')

    per_prodi = SurveiVMTS.objects.values('prodi').annotate(
        jumlah=Count('id'),
        rata=Avg('skor_total'),
        avg_v=Avg('skor_v'),
        avg_m=Avg('skor_m'),
        avg_t=Avg('skor_t'),
        avg_s=Avg('skor_s'),
    ).order_by('-rata')

    # Ambil target per prodi dari ProdiProfile
    prodi_profiles = {
        p.kode_prodi: p for p in ProdiProfile.objects.filter(aktif=True)
    }

    # Gabungkan target ke per_prodi
    from django.db import connections
    cursor = connections['default'].cursor()
    per_prodi_list = []
    for p in per_prodi:
        nama_prodi = p['prodi']
        # Cari ProdiProfile berdasarkan nama_prodi dari SIMDA
        cursor.execute("""
            SELECT m.kode_prodi 
            FROM master.program_studi ps
            JOIN mapping_prodi_instrumen m ON m.kode_prodi = ps.kode_prodi
            WHERE CONCAT(ps.nama_prodi, ' (', ps.jenjang, ')') = %s
            LIMIT 1
        """, [nama_prodi])
        row = cursor.fetchone()
        kode = row[0] if row else None
        profile = prodi_profiles.get(kode) if kode else None
        target = profile.target_survei_total if profile else site.survei_vmts_target
        pct = round((p['jumlah'] / target) * 100) if target else 0
        per_prodi_list.append({
            **p,
            'target': target,
            'pct_target': min(pct, 100),
            'target_mhs': profile.target_survei_mahasiswa if profile else '-',
            'target_dsn': profile.target_survei_dosen if profile else '-',
            'target_tdk': profile.target_survei_tendik if profile else '-',
            'target_alm': profile.target_survei_alumni if profile else '-',
        })

    daftar_prodi_filter = SurveiVMTS.objects.values_list(
        'prodi', flat=True
    ).distinct().order_by('prodi')

    pct_pemahaman = 0
    if stats['avg_total']:
        pct_pemahaman = round((stats['avg_total'] / 5) * 100, 1)

    # Data distribusi jawaban untuk grafik PIE (per pilar)
    # Ambil dari data mentah per butir — kita simpan skor bulat
    from collections import Counter

    def distribusi_skor(field):
        """Hitung distribusi skor 1-5 untuk satu pilar."""
        counts = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
        for val in data.values_list(field, flat=True):
            if val is not None:
                rounded = round(val)
                if 1 <= rounded <= 5:
                    counts[rounded] += 1
        total = sum(counts.values())
        return {
            'labels': ['1 - Sangat Tidak Setuju', '2 - Tidak Setuju', '3 - Netral', '4 - Setuju', '5 - Sangat Setuju'],
            'data': [counts[i] for i in range(1, 6)],
            'pct': [round((counts[i]/total)*100, 1) if total else 0 for i in range(1, 6)],
        }

    pie_v = distribusi_skor('skor_v')
    pie_m = distribusi_skor('skor_m')
    pie_t = distribusi_skor('skor_t')
    pie_s = distribusi_skor('skor_s')

    import json
    context = {
        'stats': stats,
        'per_status': per_status,
        'per_prodi': per_prodi_list,
        'data': data.order_by('-created_at')[:50],
        'jumlah_prodi': per_prodi.count(),
        'target_responden': site.survei_vmts_target,
        'tahun_akademik': site.survei_vmts_tahun_akademik,
        'daftar_prodi_filter': daftar_prodi_filter,
        'prodi_filter': prodi_filter,
        'pct_pemahaman': pct_pemahaman,
        'pie_v': json.dumps(pie_v),
        'pie_m': json.dumps(pie_m),
        'pie_t': json.dumps(pie_t),
        'pie_s': json.dumps(pie_s),
    }
    return render(request, 'survei/vmts_dashboard.html', context)

def dashboard_vmts_export(request):
    """Export data survei VMTS ke Excel"""
    if not request.user.is_authenticated:
        return redirect('core:login')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from django.http import HttpResponse
        return HttpResponse("openpyxl belum terinstall. Jalankan: pip install openpyxl", status=500)

    from django.http import HttpResponse
    from core.models import SiteProfile
    site = SiteProfile.get_instance()

    wb = openpyxl.Workbook()

    # Sheet 1: Data Responden
    ws1 = wb.active
    ws1.title = "Data Responden"
    header_fill = PatternFill("solid", fgColor="1a3a8f")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    headers = [
        'No', 'Tanggal', 'Nama', 'Status', 'Program Studi',
        'Angkatan', 'Skor Visi', 'Skor Misi', 'Skor Tujuan',
        'Skor Sasaran', 'Skor Total', 'Media Sosialisasi',
        'Masukan', 'Saran'
    ]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    data = SurveiVMTS.objects.all().order_by('-created_at')
    for i, r in enumerate(data, 1):
        ws1.append([
            i,
            r.created_at.strftime('%d/%m/%Y %H:%M'),
            r.nama or '-',
            r.get_status_display(),
            r.prodi,
            r.angkatan or '-',
            r.skor_v,
            r.skor_m,
            r.skor_t,
            r.skor_s,
            r.skor_total,
            r.media_sosialisasi or '-',
            r.masukan or '-',
            r.saran or '-',
        ])

    for col in ws1.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # Sheet 2: Rekap per Prodi
    ws2 = wb.create_sheet("Rekap per Prodi")
    headers2 = ['Program Studi', 'Jumlah', 'Rata-rata', 'Visi', 'Misi', 'Tujuan', 'Sasaran']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    per_prodi = SurveiVMTS.objects.values('prodi').annotate(
        jumlah=Count('id'),
        rata=Avg('skor_total'),
        avg_v=Avg('skor_v'),
        avg_m=Avg('skor_m'),
        avg_t=Avg('skor_t'),
        avg_s=Avg('skor_s'),
    ).order_by('-rata')

    for p in per_prodi:
        ws2.append([
            p['prodi'],
            p['jumlah'],
            round(p['rata'], 2) if p['rata'] else 0,
            round(p['avg_v'], 2) if p['avg_v'] else 0,
            round(p['avg_m'], 2) if p['avg_m'] else 0,
            round(p['avg_t'], 2) if p['avg_t'] else 0,
            round(p['avg_s'], 2) if p['avg_s'] else 0,
        ])

    for col in ws2.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # Sheet 3: Rekap per Status
    ws3 = wb.create_sheet("Rekap per Status")
    headers3 = ['Status', 'Jumlah', 'Rata-rata VMTS']
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    per_status = SurveiVMTS.objects.values('status').annotate(
        jumlah=Count('id'),
        rata=Avg('skor_total')
    ).order_by('status')

    STATUS_LABEL = {
        'mahasiswa': 'Mahasiswa',
        'dosen': 'Dosen',
        'tendik': 'Tenaga Kependidikan',
        'alumni': 'Alumni',
    }
    for s in per_status:
        ws3.append([
            STATUS_LABEL.get(s['status'], s['status']),
            s['jumlah'],
            round(s['rata'], 2) if s['rata'] else 0,
        ])

    # Kirim response
    filename = f"Survei_VMTS_{site.survei_vmts_tahun_akademik.replace('/', '-')}.xlsx"
    from django.http import HttpResponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response