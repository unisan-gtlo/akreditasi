"""
Excel Parser & Validator untuk Import Master Akreditasi.

Format Excel yang diharapkan:

Sheet "SubStandar":
    | nomor_standar | nomor_substandar | nomor_parent | nama | deskripsi |
    | 1             | 1.1              | (kosong)     | ...  | ...       |
    | 1             | 1.1.1            | 1.1          | ...  | ...       |

Sheet "ButirDokumen":
    | nomor_substandar | kode_butir | nama_dokumen | kategori   | wajib | format | ukuran_max | akses    | deskripsi |
    | 1.1              | 1.1-A      | ...          | UNIVERSITAS| Y     | PDF    | 50         | INTERNAL | ...       |
"""
from dataclasses import dataclass, field
from typing import List, Dict, Optional

from openpyxl import load_workbook

from .models import (
    Instrumen,
    Standar,
    SubStandar,
    ButirDokumen,
    KategoriKepemilikan,
)


# =========================================================
# DATA CLASSES
# =========================================================

@dataclass
class ParsedRow:
    """Representasi 1 baris yang sudah diparsing dari Excel."""
    row_number: int
    sheet: str
    data: dict = field(default_factory=dict)
    is_valid: bool = True
    errors: List[str] = field(default_factory=list)
    action: str = "WILL_CREATE"   # WILL_CREATE / WILL_UPDATE / DUPLICATE / INVALID
    
    def add_error(self, msg):
        self.errors.append(msg)
        self.is_valid = False
        self.action = "INVALID"


@dataclass
class ParseResult:
    """Hasil parsing 1 file Excel (2 sheet)."""
    substandar_rows: List[ParsedRow] = field(default_factory=list)
    butir_rows: List[ParsedRow] = field(default_factory=list)
    global_errors: List[str] = field(default_factory=list)
    
    @property
    def total_rows(self):
        return len(self.substandar_rows) + len(self.butir_rows)
    
    @property
    def valid_rows(self):
        return sum(1 for r in self.substandar_rows + self.butir_rows if r.is_valid)
    
    @property
    def error_rows(self):
        return self.total_rows - self.valid_rows
    
    @property
    def has_global_errors(self):
        return len(self.global_errors) > 0


# =========================================================
# EXPECTED COLUMNS
# =========================================================

SUBSTANDAR_COLUMNS = [
    "nomor_standar",
    "nomor_substandar",
    "nomor_parent",
    "nama",
    "deskripsi",
]

BUTIR_COLUMNS = [
    "nomor_substandar",
    "kode_butir",
    "nama_dokumen",
    "kategori",
    "wajib",
    "format",
    "ukuran_max",
    "akses",
    "deskripsi",
]

VALID_KATEGORI = [k[0] for k in KategoriKepemilikan.choices]
VALID_FORMAT = ["PDF", "DOCX", "XLSX", "PPTX", "GAMBAR", "APAPUN"]
VALID_AKSES = ["TERBUKA", "INTERNAL"]


# =========================================================
# PARSER
# =========================================================

class ExcelImportParser:
    """
    Parser + Validator untuk file Excel import master.
    
    Usage:
        parser = ExcelImportParser(file_path_or_object, instrumen)
        result = parser.parse_and_validate(mode="UPDATE")
        if not result.has_global_errors:
            # simpan ke ImportLog + ImportLogItem
            ...
    """
    
    def __init__(self, file_or_path, instrumen: Instrumen):
        self.file = file_or_path
        self.instrumen = instrumen
        self.result = ParseResult()
        
        # Cache existing records for fast lookup
        self._standar_cache = {}      # {nomor: Standar}
        self._substandar_cache = {}   # {nomor: SubStandar}
        self._butir_cache = {}        # {(substandar.pk, kode): ButirDokumen}
        self._load_existing()
    
    def _load_existing(self):
        """Pre-load existing data for this instrumen (speeds up validation)."""
        for s in Standar.objects.filter(instrumen=self.instrumen):
            self._standar_cache[str(s.nomor).strip()] = s
        
        for ss in SubStandar.objects.filter(standar__instrumen=self.instrumen):
            self._substandar_cache[str(ss.nomor).strip()] = ss
        
        for b in ButirDokumen.objects.filter(
            sub_standar__standar__instrumen=self.instrumen
        ).select_related("sub_standar"):
            key = (b.sub_standar.pk, str(b.kode).strip())
            self._butir_cache[key] = b
    
    # ==================== MAIN ====================
    
    def parse_and_validate(self, mode: str = "UPDATE") -> ParseResult:
        """Main entry point. Return ParseResult."""
        self.mode = mode
        
        try:
            wb = load_workbook(self.file, data_only=True)
        except Exception as e:
            self.result.global_errors.append(f"Gagal membuka file Excel: {str(e)}")
            return self.result
        
        sheet_names = wb.sheetnames
        
        # Check required sheets
        has_ss = any(s.lower() in ["substandar", "sub_standar", "sub-standar"] for s in sheet_names)
        has_b  = any(s.lower() in ["butirdokumen", "butir_dokumen", "butir-dokumen", "butir"] for s in sheet_names)
        
        if not has_ss and not has_b:
            self.result.global_errors.append(
                f"File Excel tidak punya sheet 'SubStandar' atau 'ButirDokumen'. "
                f"Sheet yang ada: {', '.join(sheet_names)}"
            )
            return self.result
        
        # Parse each sheet (case-insensitive)
        for sheet_name in sheet_names:
            name_lower = sheet_name.lower().strip()
            if name_lower in ["substandar", "sub_standar", "sub-standar"]:
                self._parse_substandar_sheet(wb[sheet_name])
            elif name_lower in ["butirdokumen", "butir_dokumen", "butir-dokumen", "butir"]:
                self._parse_butir_sheet(wb[sheet_name])
        
        # After parsing all, re-validate butir rows (their substandar might be new from Excel)
        self._revalidate_butir_references()
        
        return self.result
    
    # ==================== SUBSTANDAR ====================
    
    def _parse_substandar_sheet(self, ws):
        """Parse sheet SubStandar."""
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return  # empty or header only
        
        header = [self._normalize(c) for c in rows[0]]
        
        # Track nomor_substandar seen IN THIS EXCEL for duplicate detection
        seen_in_excel = set()
        
        for i, row in enumerate(rows[1:], start=2):  # row 2 = first data
            # Skip fully empty rows
            if not any(c is not None and str(c).strip() for c in row):
                continue
            
            row_obj = ParsedRow(row_number=i, sheet="SUBSTANDAR")
            
            # Map to dict
            data = {}
            for col_name in SUBSTANDAR_COLUMNS:
                data[col_name] = ""
            for idx, val in enumerate(row):
                if idx < len(header) and header[idx] in SUBSTANDAR_COLUMNS:
                    data[header[idx]] = self._cell_to_str(val)
            
            row_obj.data = data
            
            # === Validasi ===
            nomor_std = data.get("nomor_standar", "").strip()
            nomor_ss  = data.get("nomor_substandar", "").strip()
            nomor_par = data.get("nomor_parent", "").strip()
            nama      = data.get("nama", "").strip()
            
            if not nomor_std:
                row_obj.add_error("'nomor_standar' wajib diisi")
            elif nomor_std not in self._standar_cache:
                valid_nomors = sorted(self._standar_cache.keys())
                row_obj.add_error(
                    f"Standar nomor '{nomor_std}' tidak ditemukan di instrumen "
                    f"{self.instrumen.nama_singkat}. Nomor valid: {', '.join(valid_nomors)}"
                )
            
            if not nomor_ss:
                row_obj.add_error("'nomor_substandar' wajib diisi")
            
            if not nama:
                row_obj.add_error("'nama' sub-standar wajib diisi")
            
            # Duplicate check within this Excel
            if nomor_ss in seen_in_excel:
                row_obj.add_error(
                    f"'nomor_substandar' {nomor_ss} duplikat di baris sebelumnya dalam file Excel ini"
                )
            else:
                seen_in_excel.add(nomor_ss)
            
            # Parent check (jika ada)
            if nomor_par and nomor_par not in seen_in_excel and nomor_par not in self._substandar_cache:
                row_obj.add_error(
                    f"'nomor_parent' {nomor_par} tidak ditemukan (harus ada dulu di baris sebelumnya atau di database)"
                )
            
            # Action determination
            if row_obj.is_valid:
                existing = self._substandar_cache.get(nomor_ss)
                if existing:
                    if self.mode == "UPDATE":
                        row_obj.action = "WILL_UPDATE"
                    elif self.mode == "SKIP":
                        row_obj.action = "DUPLICATE"
                    else:  # ERROR
                        row_obj.add_error(f"SubStandar {nomor_ss} sudah ada (mode: ERROR)")
                else:
                    row_obj.action = "WILL_CREATE"
            
            self.result.substandar_rows.append(row_obj)
    
    # ==================== BUTIR ====================
    
    def _parse_butir_sheet(self, ws):
        """Parse sheet ButirDokumen."""
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return
        
        header = [self._normalize(c) for c in rows[0]]
        
        # Track (substandar, kode) seen in Excel
        seen_in_excel = set()
        
        for i, row in enumerate(rows[1:], start=2):
            if not any(c is not None and str(c).strip() for c in row):
                continue
            
            row_obj = ParsedRow(row_number=i, sheet="BUTIR")
            
            data = {}
            for col_name in BUTIR_COLUMNS:
                data[col_name] = ""
            for idx, val in enumerate(row):
                if idx < len(header) and header[idx] in BUTIR_COLUMNS:
                    data[header[idx]] = self._cell_to_str(val)
            
            row_obj.data = data
            
            # === Validasi ===
            nomor_ss = data.get("nomor_substandar", "").strip()
            kode     = data.get("kode_butir", "").strip()
            nama_dok = data.get("nama_dokumen", "").strip()
            kategori = data.get("kategori", "").strip().upper()
            wajib    = data.get("wajib", "").strip().upper()
            fmt      = data.get("format", "").strip().upper() or "PDF"
            akses    = data.get("akses", "").strip().upper() or "INTERNAL"
            
            if not nomor_ss:
                row_obj.add_error("'nomor_substandar' wajib diisi")
            # Nanti akan di-revalidate setelah semua substandar selesai di-parse
            
            if not kode:
                row_obj.add_error("'kode_butir' wajib diisi")
            
            if not nama_dok:
                row_obj.add_error("'nama_dokumen' wajib diisi")
            
            if not kategori:
                row_obj.add_error("'kategori' wajib diisi")
            elif kategori not in VALID_KATEGORI:
                row_obj.add_error(
                    f"'kategori' tidak valid. Pilihan: {', '.join(VALID_KATEGORI)}"
                )
            
            # Normalize wajib
            wajib_bool = wajib in ["Y", "YA", "YES", "TRUE", "1", "WAJIB"]
            data["_wajib_bool"] = wajib_bool
            
            if fmt and fmt not in VALID_FORMAT:
                row_obj.add_error(
                    f"'format' tidak valid. Pilihan: {', '.join(VALID_FORMAT)}"
                )
            
            if akses and akses not in VALID_AKSES:
                row_obj.add_error(
                    f"'akses' tidak valid. Pilihan: {', '.join(VALID_AKSES)}"
                )
            
            # Ukuran max
            ukuran_str = data.get("ukuran_max", "").strip() or "50"
            try:
                ukuran_int = int(float(ukuran_str))
                if ukuran_int <= 0:
                    row_obj.add_error("'ukuran_max' harus > 0")
                data["_ukuran_int"] = ukuran_int
            except (ValueError, TypeError):
                row_obj.add_error(f"'ukuran_max' harus angka, bukan '{ukuran_str}'")
                data["_ukuran_int"] = 50
            
            # Duplicate check
            excel_key = (nomor_ss, kode)
            if excel_key in seen_in_excel:
                row_obj.add_error(
                    f"Kombinasi substandar '{nomor_ss}' + kode '{kode}' duplikat di baris sebelumnya dalam file Excel ini"
                )
            else:
                seen_in_excel.add(excel_key)
            
            self.result.butir_rows.append(row_obj)
    
    def _revalidate_butir_references(self):
        """Setelah substandar selesai parsed, validasi bahwa butir.nomor_substandar ada."""
        # Kumpulkan semua nomor_substandar yang valid (existing + akan dibuat dari Excel)
        valid_substandar_nomors = set(self._substandar_cache.keys())
        for row in self.result.substandar_rows:
            if row.is_valid and row.action in ["WILL_CREATE", "WILL_UPDATE"]:
                nomor = row.data.get("nomor_substandar", "").strip()
                if nomor:
                    valid_substandar_nomors.add(nomor)
        
        for row in self.result.butir_rows:
            nomor_ss = row.data.get("nomor_substandar", "").strip()
            if nomor_ss and nomor_ss not in valid_substandar_nomors:
                row.add_error(
                    f"Sub-standar '{nomor_ss}' tidak ditemukan "
                    f"(tidak ada di database DAN tidak ada di sheet SubStandar)"
                )
            
            # Action determination untuk butir
            if row.is_valid:
                # Cari di cache (existing di database)
                # Note: kita pakai nomor_ss + kode untuk lookup, bukan pk
                existing_butir = None
                existing_ss = self._substandar_cache.get(nomor_ss)
                if existing_ss:
                    key = (existing_ss.pk, row.data.get("kode_butir", "").strip())
                    existing_butir = self._butir_cache.get(key)
                
                if existing_butir:
                    if self.mode == "UPDATE":
                        row.action = "WILL_UPDATE"
                    elif self.mode == "SKIP":
                        row.action = "DUPLICATE"
                    else:
                        row.add_error(f"Butir '{row.data['kode_butir']}' sudah ada (mode: ERROR)")
                else:
                    row.action = "WILL_CREATE"
    
    # ==================== HELPERS ====================
    
    @staticmethod
    def _normalize(val):
        """Normalize header cell: lowercase, strip, replace spaces with underscore."""
        if val is None:
            return ""
        return str(val).strip().lower().replace(" ", "_").replace("-", "_")
    
    @staticmethod
    def _cell_to_str(val):
        """Convert Excel cell to string safely."""
        if val is None:
            return ""
        if isinstance(val, float) and val.is_integer():
            return str(int(val))
        return str(val).strip()