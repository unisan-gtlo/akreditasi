"""
Generator Template Excel untuk Import Master Akreditasi.
Hasil: file .xlsx di memory (BytesIO), siap di-serve ke browser.
"""
from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .models import Instrumen, Standar


# =========================================================
# STYLING CONSTANTS
# =========================================================

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

EXAMPLE_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
EXAMPLE_FONT = Font(italic=True, color="92400E", size=10)

REF_HEADER_FILL = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
REF_HEADER_FONT = Font(bold=True, color="3730A3", size=11)

THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


# =========================================================
# MAIN GENERATOR
# =========================================================

def generate_template(instrumen: Instrumen) -> BytesIO:
    """
    Generate file Excel template untuk 1 instrumen.
    Return: BytesIO yang siap di-serve.
    """
    wb = Workbook()

    # Hapus sheet default
    wb.remove(wb.active)

    # Build 4 sheets
    _build_petunjuk_sheet(wb, instrumen)
    _build_substandar_sheet(wb, instrumen)
    _build_butir_sheet(wb, instrumen)
    _build_referensi_sheet(wb, instrumen)

    # Save ke BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# =========================================================
# SHEET 1: PETUNJUK
# =========================================================

def _build_petunjuk_sheet(wb, instrumen):
    ws = wb.create_sheet("Petunjuk", 0)
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 100

    # Title
    ws["A1"] = "📋 PETUNJUK PENGISIAN TEMPLATE IMPORT"
    ws["A1"].font = Font(bold=True, size=16, color="1E40AF")
    ws.row_dimensions[1].height = 30

    # Info instrumen
    ws["A3"] = f"Instrumen: {instrumen.nama_resmi}"
    ws["A3"].font = Font(bold=True, size=12, color="1E3A8A")
    ws["A4"] = f"Lembaga  : {instrumen.lembaga}  |  Versi: {instrumen.versi or '-'}  |  Kode: {instrumen.kode}"
    ws["A4"].font = Font(size=10, color="64748B")
    ws["A5"] = f"Di-generate: {datetime.now().strftime('%d %B %Y, %H:%M')} WITA"
    ws["A5"].font = Font(size=10, italic=True, color="94A3B8")

    # Petunjuk
    items = [
        ("", ""),
        ("CARA MENGISI", "HEADER"),
        ("", ""),
        ("1. File ini terdiri dari 4 sheet:", ""),
        ("   • Petunjuk (sheet ini) - hanya untuk informasi", ""),
        ("   • SubStandar - isi sub-kriteria per standar", ""),
        ("   • ButirDokumen - isi dokumen yang harus diarsipkan", ""),
        ("   • Referensi - daftar nilai valid (JANGAN diubah)", ""),
        ("", ""),
        ("2. Isi data mulai dari BARIS 3 pada sheet SubStandar & ButirDokumen", ""),
        ("   (Baris 1 = header, Baris 2 = contoh - BOLEH dihapus saat upload)", ""),
        ("", ""),
        ("3. Kolom WAJIB diisi: ditandai dengan warna biru di header", ""),
        ("", ""),
        ("4. Urutan pengisian SubStandar:", ""),
        ("   • Isi sub-standar level atas (tanpa parent) dulu", ""),
        ("   • Baru isi sub-standar yang punya parent", ""),
        ("   • Sistem akan resolve parent otomatis dengan multi-pass", ""),
        ("", ""),
        ("5. Kolom KATEGORI di ButirDokumen hanya boleh:", ""),
        ("   UNIVERSITAS / BIRO / FAKULTAS / PRODI", ""),
        ("", ""),
        ("6. Kolom WAJIB di ButirDokumen boleh: Y atau N", ""),
        ("", ""),
        ("7. Setelah selesai isi, upload file ini via menu Import Excel di SIAKRED", ""),
        ("", ""),
        ("", ""),
        ("TIPS", "HEADER"),
        ("", ""),
        ("✓ Jangan ubah NAMA SHEET", ""),
        ("✓ Jangan ubah BARIS HEADER (baris 1)", ""),
        ("✓ Lihat sheet Referensi untuk daftar Nomor Standar yang valid", ""),
        ("✓ Preview dulu sebelum commit - sistem akan validasi per baris", ""),
        ("", ""),
        ("", ""),
        ("KONTAK BANTUAN", "HEADER"),
        ("", ""),
        ("Pustikom UNISAN - akademik.unisan@gmail.com", ""),
    ]

    for idx, (text, style) in enumerate(items, start=7):
        cell = ws.cell(row=idx, column=1, value=text)
        if style == "HEADER":
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
            ws.row_dimensions[idx].height = 22
        else:
            cell.font = Font(size=10, color="334155")


# =========================================================
# SHEET 2: SUBSTANDAR
# =========================================================

def _build_substandar_sheet(wb, instrumen):
    ws = wb.create_sheet("SubStandar")

    # Header
    headers = [
        ("nomor_standar", 15, True),
        ("nomor_substandar", 18, True),
        ("nomor_parent", 15, False),
        ("nama", 50, True),
        ("deskripsi", 60, False),
    ]

    for col_idx, (header_text, width, required) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Contoh rows (2-3 baris)
    # Pakai standar pertama yang ada di instrumen sebagai contoh
    first_standar = Standar.objects.filter(instrumen=instrumen).order_by("urutan").first()
    if first_standar:
        std_nomor = first_standar.nomor
    else:
        std_nomor = "1"

    example_rows = [
        [std_nomor, f"{std_nomor}.1", "", "Contoh: Sub-standar level 1", "Deskripsi optional"],
        [std_nomor, f"{std_nomor}.1.1", f"{std_nomor}.1", "Contoh: Sub-sub-standar (anak dari di atas)", "Nested substandar"],
    ]

    for row_idx, row_data in enumerate(example_rows, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = EXAMPLE_FILL
            cell.font = EXAMPLE_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Data Validation: nomor_standar harus dari daftar yang valid
    valid_nomors = list(
        Standar.objects.filter(instrumen=instrumen)
        .order_by("urutan")
        .values_list("nomor", flat=True)
    )
    if valid_nomors:
        formula = '"' + ",".join(str(n) for n in valid_nomors) + '"'
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showDropDown=False,  # False = TAMPILKAN arrow
            errorTitle="Nomor Standar Invalid",
            error=f"Pilih dari: {', '.join(valid_nomors)}",
        )
        ws.add_data_validation(dv)
        # Apply ke kolom A baris 4-200 (skip baris contoh)
        dv.add("A4:A200")


# =========================================================
# SHEET 3: BUTIR DOKUMEN
# =========================================================

def _build_butir_sheet(wb, instrumen):
    ws = wb.create_sheet("ButirDokumen")

    headers = [
        ("nomor_substandar", 18, True),
        ("kode_butir", 15, True),
        ("nama_dokumen", 50, True),
        ("kategori", 15, True),
        ("wajib", 10, True),
        ("format", 12, False),
        ("ukuran_max", 12, False),
        ("akses", 12, False),
        ("deskripsi", 40, False),
    ]

    for col_idx, (header_text, width, required) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    first_standar = Standar.objects.filter(instrumen=instrumen).order_by("urutan").first()
    std_nomor = first_standar.nomor if first_standar else "1"

    # Contoh rows
    example_rows = [
        [f"{std_nomor}.1", f"{std_nomor}.1-A", "Contoh: Dokumen SK Senat tentang Statuta",
         "UNIVERSITAS", "Y", "PDF", 50, "INTERNAL", "Deskripsi optional"],
        [f"{std_nomor}.1", f"{std_nomor}.1-B", "Contoh: Renstra Prodi 2024-2029",
         "PRODI", "Y", "PDF", 30, "INTERNAL", ""],
        [f"{std_nomor}.1.1", f"{std_nomor}.1.1-A", "Contoh: SK Rektor Pengesahan VMTS (dokumen terbuka)",
         "UNIVERSITAS", "N", "PDF", 10, "TERBUKA", "Publik bisa akses"],
    ]

    for row_idx, row_data in enumerate(example_rows, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = EXAMPLE_FILL
            cell.font = EXAMPLE_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Data Validations

    # Kategori dropdown
    dv_kat = DataValidation(
        type="list",
        formula1='"UNIVERSITAS,BIRO,FAKULTAS,PRODI"',
        allow_blank=False,
        errorTitle="Kategori Invalid",
        error="Pilih: UNIVERSITAS, BIRO, FAKULTAS, atau PRODI",
    )
    ws.add_data_validation(dv_kat)
    dv_kat.add("D5:D200")

    # Wajib dropdown
    dv_wajib = DataValidation(
        type="list",
        formula1='"Y,N"',
        allow_blank=False,
        errorTitle="Wajib Invalid",
        error="Isi Y (wajib) atau N (opsional)",
    )
    ws.add_data_validation(dv_wajib)
    dv_wajib.add("E5:E200")

    # Format dropdown
    dv_fmt = DataValidation(
        type="list",
        formula1='"PDF,DOCX,XLSX,PPTX,GAMBAR,APAPUN"',
        allow_blank=True,
        errorTitle="Format Invalid",
        error="Pilih dari dropdown",
    )
    ws.add_data_validation(dv_fmt)
    dv_fmt.add("F5:F200")

    # Akses dropdown
    dv_akses = DataValidation(
        type="list",
        formula1='"TERBUKA,INTERNAL"',
        allow_blank=True,
        errorTitle="Akses Invalid",
        error="Pilih: TERBUKA atau INTERNAL",
    )
    ws.add_data_validation(dv_akses)
    dv_akses.add("H5:H200")


# =========================================================
# SHEET 4: REFERENSI
# =========================================================

def _build_referensi_sheet(wb, instrumen):
    ws = wb.create_sheet("Referensi")
    ws.sheet_view.showGridLines = False

    # Info
    ws["A1"] = f"REFERENSI - {instrumen.nama_singkat}"
    ws["A1"].font = Font(bold=True, size=14, color="1E40AF")
    ws.merge_cells("A1:D1")

    ws["A2"] = "Daftar nilai yang valid untuk pengisian. Jangan diubah."
    ws["A2"].font = Font(size=10, italic=True, color="64748B")
    ws.merge_cells("A2:D2")

    # Daftar standar per instrumen
    ws["A4"] = "NOMOR STANDAR"
    ws["B4"] = "NAMA STANDAR"
    ws["C4"] = "BOBOT (%)"

    for col in ["A4", "B4", "C4"]:
        ws[col].font = REF_HEADER_FONT
        ws[col].fill = REF_HEADER_FILL
        ws[col].alignment = HEADER_ALIGN
        ws[col].border = THIN_BORDER

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 30
    ws.row_dimensions[4].height = 24

    row_idx = 5
    for std in Standar.objects.filter(instrumen=instrumen).order_by("urutan", "nomor"):
        ws.cell(row=row_idx, column=1, value=std.nomor).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=std.nama).border = THIN_BORDER
        ws.cell(row=row_idx, column=3, value=str(std.bobot) if std.bobot else "-").border = THIN_BORDER
        row_idx += 1

    # Separator
    row_idx += 2

    # Nilai valid untuk kolom ButirDokumen
    ws.cell(row=row_idx, column=1, value="KOLOM BUTIR DOKUMEN").font = REF_HEADER_FONT
    ws.cell(row=row_idx, column=1).fill = REF_HEADER_FILL
    ws.cell(row=row_idx, column=2, value="NILAI VALID").font = REF_HEADER_FONT
    ws.cell(row=row_idx, column=2).fill = REF_HEADER_FILL
    ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=4)
    row_idx += 1

    references = [
        ("kategori", "UNIVERSITAS, BIRO, FAKULTAS, PRODI"),
        ("wajib", "Y, N"),
        ("format", "PDF, DOCX, XLSX, PPTX, GAMBAR, APAPUN"),
        ("akses", "TERBUKA, INTERNAL"),
        ("ukuran_max", "angka dalam MB (contoh: 50)"),
    ]

    for key, valid_values in references:
        ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True, size=10, color="3730A3")
        ws.cell(row=row_idx, column=1).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=valid_values).font = Font(size=10, color="334155")
        ws.cell(row=row_idx, column=2).border = THIN_BORDER
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=4)
        row_idx += 1