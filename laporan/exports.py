"""Helper functions untuk export laporan ke PDF dan Excel.

Semua fungsi menerima data yang sudah di-process dan return HttpResponse 
yang siap di-download oleh browser.
"""
from datetime import datetime
from io import BytesIO

from django.http import HttpResponse


# ============================================================
# PDF HELPERS (reportlab)
# ============================================================

def _get_pdf_base_styles():
    """Build dict of reportlab styles yang konsisten semua laporan."""
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    
    base = getSampleStyleSheet()
    styles = {
        "title": ParagraphStyle(
            "TitleSIAKRED", parent=base["Title"],
            fontSize=14, textColor=colors.HexColor("#1E3A8A"),
            spaceAfter=6, alignment=TA_CENTER, fontName="Helvetica-Bold",
        ),
        "subtitle": ParagraphStyle(
            "SubtitleSIAKRED", parent=base["Normal"],
            fontSize=11, textColor=colors.HexColor("#475569"),
            spaceAfter=12, alignment=TA_CENTER,
        ),
        "section": ParagraphStyle(
            "SectionSIAKRED", parent=base["Heading2"],
            fontSize=11, textColor=colors.HexColor("#1E40AF"),
            spaceBefore=10, spaceAfter=6, fontName="Helvetica-Bold",
        ),
        "normal": ParagraphStyle(
            "NormalSIAKRED", parent=base["Normal"],
            fontSize=9, leading=12,
        ),
        "small": ParagraphStyle(
            "SmallSIAKRED", parent=base["Normal"],
            fontSize=7, textColor=colors.HexColor("#6B7280"),
            alignment=TA_CENTER,
        ),
        "footer": ParagraphStyle(
            "FooterSIAKRED", parent=base["Normal"],
            fontSize=7, textColor=colors.HexColor("#9CA3AF"),
            alignment=TA_CENTER,
        ),
    }
    return styles


def _pdf_header(title, subtitle=None):
    """Return list of flowables untuk header PDF."""
    from reportlab.platypus import Paragraph, Spacer
    from reportlab.lib.units import mm
    
    styles = _get_pdf_base_styles()
    story = []
    
    # Header SIAKRED
    header_style = styles["small"]
    story.append(Paragraph("<b>SIAKRED</b> &middot; Sistem Informasi Akreditasi UNISAN", header_style))
    story.append(Spacer(1, 4 * mm))
    
    # Title
    story.append(Paragraph(title, styles["title"]))
    if subtitle:
        story.append(Paragraph(subtitle, styles["subtitle"]))
    
    return story


def _pdf_footer_text():
    """Footer text untuk PDF."""
    return f"Dicetak: {datetime.now().strftime('%d %B %Y %H:%M')} &middot; PUSTIKOM UNISAN &middot; SIAKRED v1.2"


# ============================================================
# PDF: LAPORAN PROGRESS PER SESI
# ============================================================

def export_sesi_pdf(sesi, data):
    """Generate PDF laporan progress sesi.
    
    Args:
        sesi: SesiAkreditasi object
        data: dict hasil _get_sesi_progress_data(sesi)
    
    Returns:
        HttpResponse dengan PDF content
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    )
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
        title=f"Laporan Progress: {sesi.judul}",
    )
    
    styles = _get_pdf_base_styles()
    story = []
    
    # HEADER
    instrumen_name = (
        getattr(sesi.instrumen, "nama", None)
        or getattr(sesi.instrumen, "judul", None)
        or getattr(sesi.instrumen, "kode", None)
        or str(sesi.instrumen)
    )
    story.extend(_pdf_header(
        "LAPORAN PROGRESS SESI AKREDITASI",
        sesi.judul,
    ))
    
    # INFO SESI
    story.append(Paragraph("INFORMASI SESI", styles["section"]))
    
    info_data = [
        ["Prodi", f"{sesi.kode_prodi or '-'} {sesi.nama_prodi_snapshot or ''}"],
        ["Instrumen", instrumen_name],
        ["Tahun TS", sesi.tahun_ts or "-"],
        ["Status", sesi.get_status_display()],
        ["Tanggal Mulai", sesi.tanggal_mulai.strftime("%d %B %Y") if sesi.tanggal_mulai else "-"],
        ["Target Selesai", sesi.tanggal_target_selesai.strftime("%d %B %Y") if sesi.tanggal_target_selesai else "-"],
    ]
    
    info_table = Table(info_data, colWidths=[45 * mm, 130 * mm])
    info_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#6B7280")),
        ("TEXTCOLOR", (1, 0), (1, -1), colors.HexColor("#111827")),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.HexColor("#E5E7EB")),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 6 * mm))
    
    # RINGKASAN STATS
    story.append(Paragraph("RINGKASAN STATUS", styles["section"]))
    
    stats = data["stats"]
    stats_header = ["Total Butir", "Terisi", "Approved", "Pending", "Revisi", "Rejected"]
    stats_values = [
        str(stats["total_butir"]),
        str(stats["butir_terisi"]),
        str(stats["status_approved"]),
        str(stats["status_pending"]),
        str(stats["status_revision"]),
        str(stats["status_rejected"]),
    ]
    
    stats_table = Table([stats_header, stats_values], colWidths=[29 * mm] * 6)
    stats_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, 0), 7),
        ("FONTSIZE", (0, 1), (-1, 1), 14),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica"),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#6B7280")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F9FAFB")),
        ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.HexColor("#E5E7EB")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E5E7EB")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        # Coloring value row
        ("TEXTCOLOR", (1, 1), (1, 1), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (2, 1), (2, 1), colors.HexColor("#059669")),
        ("TEXTCOLOR", (3, 1), (3, 1), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (4, 1), (4, 1), colors.HexColor("#D97706")),
        ("TEXTCOLOR", (5, 1), (5, 1), colors.HexColor("#DC2626")),
    ]))
    story.append(stats_table)
    story.append(Spacer(1, 4 * mm))
    
    # PROGRESS TEXT
    progress_text = (
        f"<b>Progress Kelengkapan:</b> {stats['progress_pct']}% "
        f"({stats['butir_terisi']}/{stats['total_butir']} butir terisi, "
        f"Approved: {stats['approved_pct']}%)"
    )
    story.append(Paragraph(progress_text, styles["normal"]))
    story.append(Spacer(1, 8 * mm))
    
    # DETAIL PER STANDAR
    story.append(Paragraph("DETAIL PER STANDAR", styles["section"]))
    
    for group in data["standar_groups"]:
        # Section header per Standar
        std_header = f"<b>Standar {group['nomor']}</b> &middot; {group['nama']}"
        std_meta = f"<font color='#6B7280'>{group['terisi']}/{group['total']} terisi</font>"
        story.append(Paragraph(f"{std_header} &nbsp;&nbsp;&nbsp; {std_meta}", styles["normal"]))
        story.append(Spacer(1, 2 * mm))
        
        # Table butir
        tbl_data = [["Butir", "Dokumen", "Status", "Verifikator"]]
        for item in group["butir_list"]:
            dok_text = item["dokumen"]["judul"] if item["dokumen"] else "— belum diisi —"
            status_text = {
                "APPROVED": "Approved",
                "REJECTED": "Rejected",
                "NEED_REVISION": "Need Revision",
                "PENDING": "Pending",
                "KOSONG": "Kosong",
            }.get(item["status"], item["status"])
            verif = item["verifikator"] or "-"
            tbl_data.append([
                item["kode"],
                Paragraph(dok_text[:80], styles["normal"]),
                status_text,
                verif[:25],
            ])
        
        butir_table = Table(tbl_data, colWidths=[22 * mm, 95 * mm, 28 * mm, 35 * mm], repeatRows=1)
        butir_table.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F9FAFB")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#6B7280")),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ALIGN", (2, 0), (2, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E5E7EB")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(butir_table)
        story.append(Spacer(1, 5 * mm))
    
    # FOOTER
    story.append(Spacer(1, 5 * mm))
    story.append(Paragraph("_" * 100, styles["small"]))
    story.append(Paragraph(_pdf_footer_text(), styles["footer"]))
    
    doc.build(story)
    
    pdf = buffer.getvalue()
    buffer.close()
    
    safe_judul = "".join(c if c.isalnum() or c in ("_", "-") else "_" for c in sesi.judul[:40])
    filename = f"laporan_progress_{safe_judul}_{datetime.now().strftime('%Y%m%d')}.pdf"
    
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ============================================================
# EXCEL: LAPORAN PROGRESS PER SESI
# ============================================================

def export_sesi_excel(sesi, data):
    """Generate Excel 3 sheets untuk laporan sesi."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # Common styles
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    title_font = Font(bold=True, size=14, color="1E3A8A")
    subtitle_font = Font(size=10, color="6B7280")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )
    label_fill = PatternFill("solid", fgColor="F9FAFB")
    
    instrumen_name = (
        getattr(sesi.instrumen, "nama", None)
        or getattr(sesi.instrumen, "judul", None)
        or getattr(sesi.instrumen, "kode", None)
        or str(sesi.instrumen)
    )
    
    # SHEET 1: INFO
    ws1 = wb.active
    ws1.title = "Info Sesi"
    
    ws1["A1"] = "LAPORAN PROGRESS SESI AKREDITASI"
    ws1["A1"].font = title_font
    ws1.merge_cells("A1:C1")
    
    ws1["A2"] = sesi.judul
    ws1["A2"].font = Font(bold=True, size=11)
    ws1.merge_cells("A2:C2")
    
    ws1["A3"] = f"Dicetak: {datetime.now().strftime('%d %B %Y %H:%M')}"
    ws1["A3"].font = subtitle_font
    ws1.merge_cells("A3:C3")
    
    info_rows = [
        ("Prodi", f"{sesi.kode_prodi or '-'} {sesi.nama_prodi_snapshot or ''}"),
        ("Instrumen", instrumen_name),
        ("Tahun TS", sesi.tahun_ts or "-"),
        ("Status", sesi.get_status_display()),
        ("Tanggal Mulai", sesi.tanggal_mulai.strftime("%d %B %Y") if sesi.tanggal_mulai else "-"),
        ("Target Selesai", sesi.tanggal_target_selesai.strftime("%d %B %Y") if sesi.tanggal_target_selesai else "-"),
    ]
    
    start_row = 5
    for i, (label, value) in enumerate(info_rows):
        r = start_row + i
        ws1.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=r, column=1).fill = label_fill
        ws1.cell(row=r, column=2, value=value)
        for col in range(1, 3):
            ws1.cell(row=r, column=col).border = thin_border
    
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 55
    
    # SHEET 2: RINGKASAN
    ws2 = wb.create_sheet("Ringkasan")
    ws2["A1"] = "RINGKASAN STATUS VERIFIKASI"
    ws2["A1"].font = title_font
    ws2.merge_cells("A1:B1")
    
    stats = data["stats"]
    summary_rows = [
        ("Total Butir", stats["total_butir"]),
        ("Butir Terisi", stats["butir_terisi"]),
        ("Butir Kosong", stats["butir_kosong"]),
        ("Status: Approved", stats["status_approved"]),
        ("Status: Pending", stats["status_pending"]),
        ("Status: Need Revision", stats["status_revision"]),
        ("Status: Rejected", stats["status_rejected"]),
        ("Progress Kelengkapan (%)", f"{stats['progress_pct']}%"),
        ("Approved Rate (%)", f"{stats['approved_pct']}%"),
    ]
    
    start_row = 3
    for i, (label, value) in enumerate(summary_rows):
        r = start_row + i
        ws2.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=r, column=1).fill = label_fill
        ws2.cell(row=r, column=2, value=value)
        ws2.cell(row=r, column=2).alignment = Alignment(horizontal="right")
        for col in range(1, 3):
            ws2.cell(row=r, column=col).border = thin_border
    
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20
    
    # SHEET 3: DETAIL BUTIR
    ws3 = wb.create_sheet("Detail Butir")
    ws3["A1"] = "DETAIL PER BUTIR DOKUMEN"
    ws3["A1"].font = title_font
    ws3.merge_cells("A1:G1")
    
    headers = ["No", "Standar", "Kode Butir", "Nama Butir", "Dokumen", "Status", "Verifikator"]
    for i, h in enumerate(headers, 1):
        cell = ws3.cell(row=3, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    row_num = 4
    for idx, item in enumerate(data["butir_list"], 1):
        dok_text = item["dokumen"]["judul"] if item["dokumen"] else ""
        status_text = {
            "APPROVED": "Approved",
            "REJECTED": "Rejected",
            "NEED_REVISION": "Need Revision",
            "PENDING": "Pending",
            "KOSONG": "Kosong",
        }.get(item["status"], item["status"])
        
        row_data = [
            idx,
            f"Std {item['standar_nomor']} - {item['standar_nama'][:40]}",
            item["kode"],
            item["nama"],
            dok_text,
            status_text,
            item["verifikator"] or "",
        ]
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws3.cell(row=row_num, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # Color status cell
        status_cell = ws3.cell(row=row_num, column=6)
        if item["status"] == "APPROVED":
            status_cell.fill = PatternFill("solid", fgColor="D1FAE5")
        elif item["status"] == "REJECTED":
            status_cell.fill = PatternFill("solid", fgColor="FEE2E2")
        elif item["status"] == "NEED_REVISION":
            status_cell.fill = PatternFill("solid", fgColor="FEF3C7")
        elif item["status"] == "PENDING":
            status_cell.fill = PatternFill("solid", fgColor="DBEAFE")
        
        row_num += 1
    
    # Column widths
    col_widths = [5, 30, 14, 40, 50, 16, 20]
    for i, w in enumerate(col_widths, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    
    # Freeze header
    ws3.freeze_panes = "A4"
    
    # Save
    buffer = BytesIO()
    wb.save(buffer)
    xlsx = buffer.getvalue()
    buffer.close()
    
    safe_judul = "".join(c if c.isalnum() or c in ("_", "-") else "_" for c in sesi.judul[:40])
    filename = f"laporan_progress_{safe_judul}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    response = HttpResponse(
        xlsx,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ============================================================
# EXCEL: COMPLETENESS PRODI
# ============================================================

def export_prodi_excel(prodi_list, agg, fakultas_filter_nama=None):
    """Generate Excel tabel ranking prodi."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Ranking Prodi"
    
    title_font = Font(bold=True, size=14, color="1E3A8A")
    subtitle_font = Font(size=10, color="6B7280")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="2563EB")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )
    
    # Header
    ws["A1"] = "LAPORAN COMPLETENESS PER PRODI"
    ws["A1"].font = title_font
    ws.merge_cells("A1:J1")
    
    subtitle = f"Dicetak: {datetime.now().strftime('%d %B %Y %H:%M')}"
    if fakultas_filter_nama:
        subtitle += f" &middot; Filter: {fakultas_filter_nama}"
    ws["A2"] = subtitle
    ws["A2"].font = subtitle_font
    ws.merge_cells("A2:J2")
    
    # Aggregate info
    ws["A4"] = f"Total Prodi: {agg['total_prodi']}"
    ws["C4"] = f"Total Butir: {agg['total_butir']}"
    ws["E4"] = f"Butir Terisi: {agg['total_terisi']}"
    ws["G4"] = f"Progress Keseluruhan: {agg['overall_progress']}%"
    for col in ["A", "C", "E", "G"]:
        ws[f"{col}4"].font = Font(bold=True, size=10)
    
    # Table Header
    headers = [
        "Rank", "Kode", "Nama Prodi", "Fakultas", "Instrumen",
        "Total", "Terisi", "Approved", "Pending", "Progress %"
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    for idx, p in enumerate(prodi_list, 1):
        row_num = 6 + idx
        row_data = [
            idx,
            p["kode_prodi"],
            p["nama_prodi"],
            f"{p['kode_fakultas']} - {p['nama_fakultas']}" if p.get("nama_fakultas") and p["nama_fakultas"] != "-" else p["kode_fakultas"],
            p["instrumen_nama"],
            p["total_butir"],
            p["butir_terisi"],
            p["approved_count"],
            p["pending_count"],
            p["progress_pct"],
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx in (1, 2, 6, 7, 8, 9, 10):
                cell.alignment = Alignment(horizontal="center")
        
        # Color progress cell by tier
        prog_cell = ws.cell(row=row_num, column=10)
        if p["tier"] == "high":
            prog_cell.fill = PatternFill("solid", fgColor="D1FAE5")
            prog_cell.font = Font(bold=True, color="065F46")
        elif p["tier"] == "mid":
            prog_cell.fill = PatternFill("solid", fgColor="FEF3C7")
            prog_cell.font = Font(bold=True, color="92400E")
        elif p["tier"] == "low":
            prog_cell.fill = PatternFill("solid", fgColor="FEE2E2")
            prog_cell.font = Font(bold=True, color="991B1B")
    
    col_widths = [7, 8, 28, 25, 24, 8, 8, 11, 10, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    ws.freeze_panes = "A7"
    
    # Save
    buffer = BytesIO()
    wb.save(buffer)
    xlsx = buffer.getvalue()
    buffer.close()
    
    filename = f"laporan_completeness_prodi_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        xlsx,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ============================================================
# PDF: HEATMAP VERIFIKASI
# ============================================================

def export_heatmap_pdf(selected_instrumen, standar_list, prodi_list, matrix, insights):
    """Generate PDF heatmap landscape."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    )
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title="Heatmap Verifikasi",
    )
    
    styles = _get_pdf_base_styles()
    story = []
    
    instrumen_name = (
        getattr(selected_instrumen, "nama", None)
        or getattr(selected_instrumen, "judul", None)
        or getattr(selected_instrumen, "kode", None)
        or "Instrumen"
    ) if selected_instrumen else "Instrumen"
    
    story.extend(_pdf_header(
        "HEATMAP VERIFIKASI DOKUMEN",
        f"Instrumen: {instrumen_name}",
    ))
    
    # MATRIX TABLE
    if standar_list and prodi_list:
        # Header row
        header = ["Standar"] + [p["kode_prodi"] for p in prodi_list]
        matrix_data = [header]
        
        # Color map
        color_map = {
            "green": colors.HexColor("#10B981"),
            "yellow": colors.HexColor("#FBBF24"),
            "blue": colors.HexColor("#60A5FA"),
            "orange": colors.HexColor("#F97316"),
            "red": colors.HexColor("#EF4444"),
            "empty": colors.HexColor("#E5E7EB"),
        }
        
        cell_styles = []
        
        for r_idx, s in enumerate(standar_list, 1):
            row = [f"Std {s['nomor']}"]
            for c_idx, p in enumerate(prodi_list, 1):
                cell = matrix.get(s["id"], {}).get(p["kode_prodi"], {})
                if cell.get("terisi"):
                    text = f"{cell['terisi']}/{cell['total']}"
                else:
                    text = "—"
                row.append(text)
                
                # Color cell
                color_key = cell.get("color", "empty")
                cell_styles.append(("BACKGROUND", (c_idx, r_idx), (c_idx, r_idx), color_map[color_key]))
                if color_key == "empty":
                    cell_styles.append(("TEXTCOLOR", (c_idx, r_idx), (c_idx, r_idx), colors.HexColor("#9CA3AF")))
                elif color_key == "yellow":
                    cell_styles.append(("TEXTCOLOR", (c_idx, r_idx), (c_idx, r_idx), colors.HexColor("#78350F")))
                else:
                    cell_styles.append(("TEXTCOLOR", (c_idx, r_idx), (c_idx, r_idx), colors.white))
            
            matrix_data.append(row)
        
        # Column widths
        num_cols = len(prodi_list) + 1
        first_col_w = 40 * mm
        remaining_w = 260 * mm - first_col_w
        other_col_w = remaining_w / len(prodi_list) if prodi_list else 20 * mm
        
        col_widths = [first_col_w] + [other_col_w] * len(prodi_list)
        
        matrix_table = Table(matrix_data, colWidths=col_widths, repeatRows=1)
        
        base_style = [
            ("FONTSIZE", (0, 0), (-1, 0), 7),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F9FAFB")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#6B7280")),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E5E7EB")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("ALIGN", (0, 1), (0, -1), "LEFT"),
            ("LEFTPADDING", (0, 1), (0, -1), 5),
        ]
        
        matrix_table.setStyle(TableStyle(base_style + cell_styles))
        story.append(matrix_table)
        story.append(Spacer(1, 4 * mm))
        
        # LEGEND
        legend_data = [[
            "Hijau: Semua Approved",
            "Kuning: Sebagian Approved",
            "Biru: Pending Review",
            "Oranye: Perlu Revisi",
            "Merah: Ada Rejected",
            "Abu: Belum Diisi",
        ]]
        legend_style = [
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("TEXTCOLOR", (0, 0), (0, 0), colors.HexColor("#065F46")),
            ("TEXTCOLOR", (1, 0), (1, 0), colors.HexColor("#92400E")),
            ("TEXTCOLOR", (2, 0), (2, 0), colors.HexColor("#1E40AF")),
            ("TEXTCOLOR", (3, 0), (3, 0), colors.HexColor("#C2410C")),
            ("TEXTCOLOR", (4, 0), (4, 0), colors.HexColor("#991B1B")),
            ("TEXTCOLOR", (5, 0), (5, 0), colors.HexColor("#6B7280")),
        ]
        legend_table = Table(legend_data, colWidths=[45 * mm] * 6)
        legend_table.setStyle(TableStyle(legend_style))
        story.append(legend_table)
    else:
        story.append(Paragraph("Tidak ada data untuk ditampilkan.", styles["normal"]))
    
    # INSIGHTS
    if insights:
        story.append(Spacer(1, 6 * mm))
        story.append(Paragraph("INSIGHTS &amp; REKOMENDASI", styles["section"]))
        for i, insight in enumerate(insights, 1):
            story.append(Paragraph(f"{i}. {insight}", styles["normal"]))
    
    # FOOTER
    story.append(Spacer(1, 5 * mm))
    story.append(Paragraph(_pdf_footer_text(), styles["footer"]))
    
    doc.build(story)
    
    pdf = buffer.getvalue()
    buffer.close()
    
    filename = f"heatmap_verifikasi_{instrumen_name[:20]}_{datetime.now().strftime('%Y%m%d')}.pdf"
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ============================================================
# EXCEL: AUDIT TRAIL
# ============================================================

def export_audit_excel(logs, stats, filters):
    """Generate Excel audit trail."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Trail"
    
    title_font = Font(bold=True, size=14, color="1E3A8A")
    subtitle_font = Font(size=10, color="6B7280")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="475569")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )
    
    ws["A1"] = "AUDIT TRAIL VERIFIKASI DOKUMEN"
    ws["A1"].font = title_font
    ws.merge_cells("A1:G1")
    
    ws["A2"] = f"Dicetak: {datetime.now().strftime('%d %B %Y %H:%M')}"
    ws["A2"].font = subtitle_font
    ws.merge_cells("A2:G2")
    
    # Filter info
    filter_text = []
    if filters.get("date_from"):
        filter_text.append(f"Dari: {filters['date_from']}")
    if filters.get("date_to"):
        filter_text.append(f"Sampai: {filters['date_to']}")
    if filters.get("aksi"):
        filter_text.append(f"Aksi: {filters['aksi']}")
    if filters.get("search"):
        filter_text.append(f"Cari: '{filters['search']}'")
    
    ws["A3"] = "Filter: " + (" | ".join(filter_text) if filter_text else "Tidak ada")
    ws["A3"].font = subtitle_font
    ws.merge_cells("A3:G3")
    
    # Stats
    ws["A5"] = f"Total: {stats['total']}"
    ws["B5"] = f"Approved: {stats['approved']}"
    ws["C5"] = f"Rejected: {stats['rejected']}"
    ws["D5"] = f"Revision: {stats['revision']}"
    ws["E5"] = f"Reset: {stats['reset']}"
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}5"].font = Font(bold=True, size=10)
    
    # Headers
    headers = [
        "Tanggal", "Waktu", "Aksi", "Status Lama", "Status Baru",
        "Dokumen", "Butir", "Prodi", "Verifikator", "Catatan"
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=7, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    for idx, log in enumerate(logs, 1):
        row_num = 7 + idx
        
        dokumen = log.verifikasi.revisi.dokumen if log.verifikasi and log.verifikasi.revisi else None
        butir_kode = dokumen.butir_dokumen.kode if dokumen and dokumen.butir_dokumen else ""
        prodi = dokumen.scope_kode_prodi if dokumen else ""
        judul_dokumen = dokumen.judul if dokumen else "—"
        
        verif_name = (
            log.dilakukan_oleh.get_full_name() or log.dilakukan_oleh.username
            if log.dilakukan_oleh else "(sistem)"
        )
        
        row_data = [
            log.tanggal.strftime("%Y-%m-%d"),
            log.tanggal.strftime("%H:%M:%S"),
            log.aksi,
            log.status_lama or "",
            log.status_baru or "",
            judul_dokumen,
            butir_kode,
            prodi or "",
            verif_name,
            log.catatan or "",
        ]
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # Color aksi cell
        aksi_cell = ws.cell(row=row_num, column=3)
        aksi_cell.alignment = Alignment(horizontal="center", vertical="center")
        aksi_cell.font = Font(bold=True)
        if log.aksi == "APPROVED":
            aksi_cell.fill = PatternFill("solid", fgColor="D1FAE5")
        elif log.aksi == "REJECTED":
            aksi_cell.fill = PatternFill("solid", fgColor="FEE2E2")
        elif log.aksi == "NEED_REVISION":
            aksi_cell.fill = PatternFill("solid", fgColor="FEF3C7")
        elif log.aksi == "RESET":
            aksi_cell.fill = PatternFill("solid", fgColor="F3F4F6")
    
    col_widths = [12, 10, 14, 14, 14, 42, 14, 10, 20, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    ws.freeze_panes = "A8"
    
    buffer = BytesIO()
    wb.save(buffer)
    xlsx = buffer.getvalue()
    buffer.close()
    
    filename = f"audit_trail_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        xlsx,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
